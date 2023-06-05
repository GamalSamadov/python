"""
The load_workbook() function enables you to open an excel file.
The sheetnames property is used to find the names of sheets in an Excel file.
You can access and edit the sheet name through the title property.
The active property lets you know the name of the active sheet.
The cell can be read through the value property.
The row property is used to find the line of the cell and the column property to find the column of the cell.
The cell can also be read through the cell() function.
The max_row property is used to find the number of sheet lines and the max_column property to find the number of sheet columns.
"""

import openpyxl
from pathlib import Path

pth = Path.home() / Path('OneDrive', 'Desktop', 'Curces', 'Hsoub Academy', 'The Code',
                         '02 Practical applications using Python', '06 Working with Excel files', 'example.xlsx')

exelfile = openpyxl.load_workbook(pth)

print(exelfile.sheetnames)

sheet1 = exelfile['Sheet1']
print(sheet1.title)

activaSheet = exelfile.active
print(activaSheet.title)

print(sheet1['A1'].value)
print(sheet1['B1'].value)
print(sheet1['C1'].value)
print(sheet1['C1'].row)
print(sheet1['C1'].column)
print(sheet1['C1'].coordinate)

print(sheet1.cell(row=1, column=2).value)

print('-' * 50)

for i in range(1, 7):
    print(i, sheet1.cell(row=i, column=1).value)

print('-' * 50)

total = 0
for i in range(1, sheet1.max_row+1):
    print(i, sheet1.cell(row=i, column=1).value, sheet1.cell(row=i, column=2).value)
    total += sheet1.cell(row=i, column=2).value

print(f'The Total Salary of the employees is {total}')

print(sheet1.max_row)
print(sheet1.max_column)