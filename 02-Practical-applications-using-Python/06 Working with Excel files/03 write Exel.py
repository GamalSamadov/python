"""
We use Workbook() builder to create a container for all parts of the Excel file.
You can create a new sheet with the create_chartsheet() function.
The save() function enables you to save an excel file.
"""
import openpyxl
from pathlib import Path

# Create excel file
exelfile = openpyxl.Workbook()

# change sheet name
exelsheet = exelfile.active
exelsheet.title = 'firstsheet'

# create sheet
exelfile.create_sheet()
exelfile.create_sheet()
exelfile.create_sheet(index=1, title='secondsheet')
exelfile.create_sheet(index=2, title='middlesheet')

# delete sheet
del exelfile['middlesheet']

# write to sheet
sheet1 = exelfile['secondsheet']
sheet1['A1'] = 'Hello world!'

# write to excel proctes
names = ['Hadi', 'Yara', 'Sara', 'Anas']
firstsheet = exelfile['firstsheet']

for i in range(1, len(names) + 1):
    firstsheet.cell(row=i, column=3).value = names[i-1]

# save exel file
exelfile.save(filename=Path.home()/Path('OneDrive', 'Desktop', 'Curces', 'Hsoub Academy', 'The Code', '02 Practical applications using Python', '06 Working with Excel files', 'NewExelSheet.xlsx'))