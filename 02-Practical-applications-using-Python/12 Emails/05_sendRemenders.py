import openpyxl, smtplib
from pathlib import Path

# Connect with Excel file
file = openpyxl.load_workbook(Path.home() / Path('OneDrive', 'Desktop', 'members.xlsx'))
sheets = file.sheetnames
sheet = file['Sheet1']
last_col = sheet.max_column
lates_month = sheet.cell(row=1, column=last_col).value
# print(lates_month)

unpaid_members = {}

for r in range(2, sheet.max_row + 1):
    payment = sheet.cell(row=r, column=last_col).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaid_members[name] = email

# Send mail
smtpObject = smtplib.SMTP('smtp.gmail.com', 587)
smtpObject.starttls()
sender_email = 'uvaysalmohammady@gmail.com'
password = 'eesahewtktbltdmd'

smtpObject.login(sender_email, password)

for name, email in unpaid_members.items():
    body = """Subject: %s dues unpaid.\nDear %s, \nRecords show that you not paid dues for %s.
    Please make this payment as soon as possible. Thank you!""" %(lates_month, name, lates_month)
    print('Sending email to %s...' % email)
    sendMailStatus = smtpObject.sendmail(sender_email, email, body)

    if sendMailStatus != {}:
        print('There was a problem sending email to %s: %s' % (email, sendMailStatus))
smtpObject.quit()